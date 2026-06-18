"""
Core processing logic for TicketVault monthly reports.
Separated from the Streamlit UI so it can be used independently or tested.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAMED_CUSTOMERS = [
    "TicketsNow",
    "Gametime",
    "Vivid Seats",
    "StubHub",
    "SeatGeek",
    "GoTickets",
    "TicketNetwork",
    "TickPick",
    "Ticket Evolution",
    "Mercury",
]

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

HEADERS = [
    "Seller Email",
    "Performer",
    "Event",
    "Event Date",
    "Venue",
    "Customer",
    "Ticket Sales",
    "QTY",
    "Invoice Date",
]

COL_WIDTHS = [18, 22, 22, 14, 30, 18, 14, 8, 14]

# ── Hardcoded company → account mapping (from Master_Mapping_List.xlsx, cols B & E) ──

COMPANY_MAPPING: dict[str, str] = {
    "damon and crew":        "ystickets",
    "bearhawk - aaron":      "ystickets",
    "bearhawk - ari":        "ystickets",
    "bearhawk - chris":      "ystickets",
    "bearhawk - dylan":      "ystickets",
    "bearhawk - gray":       "ystickets",
    "bearhawk - jason":      "ystickets",
    "bearhawk group":        "ystickets",
    "best tix":              "ystickets",
    "ticketwonders llc":     "ystickets",
    "ticketwonders2 llc":    "ystickets",
    "upside llc":            "ystickets",
    "not found":             "ystickets",
    "ys tickets":            "ystickets",
    "ys-seatgeek2":          "ystickets",
    "ys-seatgeek":           "ystickets",
    "ys tickets spec":       "ystickets",
    "yourtickets":           "ystickets",
    "your tickets":          "ystickets",
    "ysa":                   "ystickets",
    "ysa 2":                 "ystickets",
    "ysa 3":                 "ystickets",
    "jacks ys":              "ystickets",
    "ys katz":               "ystickets",
    "yoni levine":           "ystickets",
    "levovitz":              "ystickets",
    "needle tickets llc":    "ystickets",
    "ys tl":                 "ystickets",
    "gk llc":                "ystickets",
    "slash financial inc":   "ystickets",
    "ysm tickets":           "ystickets",
    "pollak tickets":        "ystickets",
    "yss tickets":           "ystickets",
    "ysw":                   "ystickets",
    "david mansbach":        "yitzknopf",
    "gra investments":       "yitzknopf",
    "gra investments - brian": "yitzknopf",
    "indiana promotions":    "yitzknopf",
    "isaac knopf":           "yitzknopf",
    "jl":                    "yitzknopf",
    "stuart levy":           "yitzknopf",
    "tv test company":       "yitzknopf",
    "the ticket guy":        "yitzknopf",
}


# ── Account lookup ────────────────────────────────────────────────────────────

def get_account(company: str) -> str:
    """
    Resolve a company name to an account using the hardcoded mapping.
    Falls back to prefix matching for variants like 'Bearhawk - Dylan'
    matching 'Bearhawk Group'.
    """
    if not company or pd.isna(company):
        return ""
    cl = str(company).strip().lower()
    if cl in COMPANY_MAPPING:
        return COMPANY_MAPPING[cl]
    for key, val in COMPANY_MAPPING.items():
        first_word = key.split(" ")[0]
        if len(first_word) > 2 and cl.startswith(first_word):
            return val
    return ""


# ── Date helpers ──────────────────────────────────────────────────────────────

def _fmt_date(d) -> str:
    if pd.isna(d):
        return ""
    try:
        return pd.to_datetime(d).strftime("%-m/%-d/%Y")
    except Exception:
        return ""


def _get_month_label(df: pd.DataFrame) -> str:
    dates = pd.to_datetime(df["Created Date"], errors="coerce")
    mode = dates.dt.to_period("M").mode()
    if mode.empty:
        return "Unknown"
    p = mode[0]
    return f"{MONTHS[p.month - 1]} {p.year}"


# ── Processing ────────────────────────────────────────────────────────────────

def process_sales(sales_files) -> dict:
    """
    Combine, deduplicate, filter, and map one or more sales export files.

    Args:
        sales_files: List of file-like objects (.xlsx TicketVault exports).

    Returns:
        dict with keys:
            month_label, ys_rows, yitz_rows, unmapped,
            total_raw, total_deduped, total_filtered
    """
    frames = []
    for f in sales_files:
        frames.append(pd.read_excel(f))
    df = pd.concat(frames, ignore_index=True)
    total_raw = len(df)

    # Deduplicate by Invoice # + Performer/Team + Event Date + Section + Row + Seats
    df["_key"] = (
        df["Invoice #"].astype(str) + "|" +
        df["Performer/Team"].astype(str) + "|" +
        df["Event Date"].astype(str) + "|" +
        df["Section"].astype(str) + "|" +
        df["Row"].astype(str) + "|" +
        df["Seats"].astype(str)
    )
    df = df.drop_duplicates(subset="_key").drop(columns="_key")
    total_deduped = len(df)

    # Filter zero-dollar rows
    df = df[pd.to_numeric(df["Total Price"], errors="coerce").fillna(0) > 0]
    total_filtered = len(df)

    month_label = _get_month_label(df)

    unmapped = set()
    ys_rows, yitz_rows = [], []

    for _, r in df.iterrows():
        account = get_account(r.get("Company"))
        if not account and r.get("Company"):
            unmapped.add(str(r["Company"]))

        ed  = pd.to_datetime(r["Event Date"],   errors="coerce")
        id_ = pd.to_datetime(r["Created Date"], errors="coerce")

        row = {
            "Seller Email": account,
            "Performer":    str(r.get("Performer/Team", "") or ""),
            "Event":        str(r.get("Performer/Opponent", "") or ""),
            "Event Date":   _fmt_date(ed),
            "Venue":        str(r.get("Venue", "") or ""),
            "Customer":     r["Client"] if r["Client"] in NAMED_CUSTOMERS else "Offsite",
            "Ticket Sales": float(r["Total Price"]),
            "QTY":          r["Quantity"],
            "Invoice Date": _fmt_date(id_),
            "_ts":          id_ if pd.notna(id_) else pd.Timestamp.min,
        }

        if account == "ystickets":
            ys_rows.append(row)
        elif account == "yitzknopf":
            yitz_rows.append(row)

    # Sort newest invoice date first
    ys_rows.sort(key=lambda r: r["_ts"], reverse=True)
    yitz_rows.sort(key=lambda r: r["_ts"], reverse=True)
    for r in ys_rows + yitz_rows:
        del r["_ts"]

    return {
        "month_label":    month_label,
        "ys_rows":        ys_rows,
        "yitz_rows":      yitz_rows,
        "unmapped":       unmapped,
        "total_raw":      total_raw,
        "total_deduped":  total_deduped,
        "total_filtered": total_filtered,
    }


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_excel(rows: list[dict], dest) -> None:
    """
    Write a list of row dicts to a styled Excel file.

    Args:
        rows: List of dicts with keys matching HEADERS.
        dest: File path or BytesIO buffer.
    """
    thin = Side(style="thin", color="D0D0D0")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Data"

    hfill = PatternFill("solid", start_color="1F3864")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 20

    bfont = Font(name="Calibri", size=11)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[h])
            cell.font = bfont
            cell.alignment = center
            cell.border = border
            if h == "Ticket Sales":
                cell.number_format = "$#,##0.00"

    for col_idx, width in zip(range(1, len(HEADERS) + 1), COL_WIDTHS):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(dest)
